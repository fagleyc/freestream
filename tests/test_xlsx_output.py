"""Tests for the Excel .xlsx output format (freestream.xlsx_writer).

Contract:

* Hdf5Recorder(output_format="xlsx") writes ONE primary ``.xlsx`` per
  point (no .h5/.mat), same ``run_NNNN_…`` basename.
* Workbook layout: one Data sheet per group (Time column + one column
  per channel, names row 1 / units row 2, numbers as NUMBERS), plus
  Meta (every root attr), Channels (waveform attrs), Devices (cal
  POINTERS) and Config (snapshot JSON) sheets.
* Sheet names are xlsx-legal (≤31 chars, sanitized, collide-safe).
* A sim sweep with each selectable format produces exactly one file of
  the selected type per point (end-to-end, fakes).
"""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook  # noqa: E402

from freestream.recorder import Hdf5Recorder  # noqa: E402
from freestream.xlsx_writer import sheet_name  # noqa: E402

RATE_SB, RATE_DB, RATE_POS = 1000.0, 100.0, 50.0
RATES = {"StrainBook_0": RATE_SB, "DaqBook2005": RATE_DB,
         "Positioner": RATE_POS}
T_START = datetime(2026, 7, 7, 14, 30, 0)

UNITS = {"StrainBook_0": {ch: "V" for ch in
                          ("N1", "N2", "Y1", "Y2", "Axial", "Roll",
                           "Excitation")},
         "DaqBook2005": {"Pdiff": "V", "Ptot": "V", "Temp": "V"},
         "Positioner": {"Alpha": "deg", "Beta": "deg"}}

DEVICES = [
    {"id": "strainbook", "model": "StrainBook/616", "sim": True,
     "cal_file": "C:/cals/sb616_2026.cal"},
    {"id": "daqbook", "model": "DaqBook/2000", "sim": False,
     "cal_file": "C:/cals/db2000.pcf"},
]

CONFIG_SNAP = {"balance": "force", "rates": {"StrainBook_0": 1000},
               "cal_pointer": "cals/x.cal"}


def make_blocks(n_sb=2000, n_db=200, n_pos=100):
    """~2000-sample balance point — the realistic per-point size."""
    rng = np.random.default_rng(42)
    return {
        "StrainBook_0": {ch: rng.normal(size=n_sb) for ch in
                         ("N1", "N2", "Y1", "Y2", "Axial", "Roll",
                          "Excitation")},
        "DaqBook2005": {ch: rng.normal(size=n_db) for ch in
                        ("Pdiff", "Ptot", "Temp")},
        "Positioner": {"Alpha": np.full(n_pos, -2.0),
                       "Beta": np.zeros(n_pos)},
    }


def write_default(rec, **kw):
    args = dict(
        point_meta={"alpha": -2.0, "beta": 0.0, "mach": 0.3,
                    "t_start": T_START, "flap_deg": 12.5},  # custom column
        blocks=make_blocks(), rates=RATES, channel_units=UNITS,
        device_meta=DEVICES, config_snapshot=CONFIG_SNAP,
    )
    args.update(kw)
    return rec.write_point(**args)


@pytest.fixture
def rec_xlsx(tmp_path):
    return Hdf5Recorder(tmp_path, config_name="cfgX",
                        output_format="xlsx")


@pytest.fixture
def workbook(rec_xlsx):
    """One written point: (path, loaded workbook, original blocks)."""
    blocks = make_blocks()
    path = write_default(rec_xlsx, blocks=blocks)
    return path, load_workbook(path), blocks


def _column(ws, col_idx, n):
    """Data column values (rows 3..3+n) as a float array."""
    return np.array([ws.cell(row=3 + i, column=col_idx).value
                     for i in range(n)], dtype=np.float64)


# ── format behavior + filename ───────────────────────────────────────────
def test_format_xlsx_writes_only_xlsx(rec_xlsx):
    p = write_default(rec_xlsx)
    assert p.suffix == ".xlsx" and p.exists()
    assert list(rec_xlsx.root_dir.rglob("*.h5")) == []
    assert list(rec_xlsx.root_dir.rglob("*.mat")) == []
    assert rec_xlsx.last_mat_path is None


def test_xlsx_filename_pattern(rec_xlsx):
    p = write_default(rec_xlsx)
    assert p.name == "run_0001_alpha_-2.0_beta_0.0_mach_0.30.xlsx"
    p2 = write_default(rec_xlsx)
    assert p2.name.startswith("run_0002_")     # numbering scans .xlsx


# ── Data sheets ──────────────────────────────────────────────────────────
def test_one_data_sheet_per_group_plus_fixed_sheets(workbook):
    _, wb, blocks = workbook
    for group in blocks:
        assert group in wb.sheetnames, f"missing Data sheet {group}"
    for fixed in ("Meta", "Channels", "Devices", "Config"):
        assert fixed in wb.sheetnames, f"missing {fixed} sheet"


def test_data_sheet_headers_units_and_time_column(workbook):
    _, wb, blocks = workbook
    ws = wb["StrainBook_0"]
    header = [c.value for c in ws[1]]
    assert header == ["Time"] + list(blocks["StrainBook_0"])
    units_row = [c.value for c in ws[2]]
    assert units_row[0] == "s"
    assert units_row[1:] == ["V"] * 7
    # Time column built from the GROUP's rate
    t = _column(ws, 1, 5)
    np.testing.assert_allclose(t, np.arange(5) / RATE_SB, rtol=0,
                               atol=1e-12)
    # a slower group gets ITS OWN rate
    ws_pos = wb["Positioner"]
    t_pos = _column(ws_pos, 1, 3)
    np.testing.assert_allclose(t_pos, np.arange(3) / RATE_POS, rtol=0,
                               atol=1e-12)


def test_channel_columns_round_trip_arrays(workbook):
    _, wb, blocks = workbook
    for group, channels in blocks.items():
        ws = wb[group]
        header = [c.value for c in ws[1]]
        for ch, arr in channels.items():
            col = header.index(ch) + 1
            got = _column(ws, col, arr.size)
            np.testing.assert_allclose(got, arr, rtol=1e-12, atol=0,
                                       err_msg=f"{group}/{ch}")
            # exactly the array's length — no stray rows below
            assert ws.cell(row=3 + arr.size, column=col).value is None


def test_numbers_are_numbers_not_strings(workbook):
    _, wb, _ = workbook
    cell = wb["StrainBook_0"].cell(row=3, column=2)
    assert isinstance(cell.value, float)
    meta = wb["Meta"]
    rows = {r[0].value: r[1].value for r in meta.iter_rows(min_row=2)}
    # xlsx numbers have no int/float distinction (-2.0 stores as -2) —
    # the requirement is numeric, NOT a string
    assert isinstance(rows["alpha"], (int, float))
    assert rows["alpha"] == -2.0
    assert isinstance(rows["flap_deg"], float)     # non-integral stays float
    assert isinstance(rows["run_number"], int)


# ── Meta sheet ───────────────────────────────────────────────────────────
def test_meta_sheet_carries_run_params(workbook):
    _, wb, _ = workbook
    rows = {r[0].value: r[1].value
            for r in wb["Meta"].iter_rows(min_row=2)}
    assert rows["run_number"] == 1
    assert rows["config_name"] == "cfgX"
    assert rows["air_state"] == "AirOn"
    assert rows["alpha"] == -2.0
    assert rows["mach"] == 0.3
    assert rows["flap_deg"] == 12.5               # custom run-sheet column
    assert rows["t_start"] == T_START.isoformat()
    assert "timestamp" in rows


# ── Channels sheet ───────────────────────────────────────────────────────
def test_channels_sheet_waveform_attrs(workbook):
    _, wb, blocks = workbook
    ws = wb["Channels"]
    header = [c.value for c in ws[1]]
    assert header == ["group", "channel", "unit", "wf_increment",
                      "wf_samples", "wf_start_time"]
    rows = {(r[0].value, r[1].value): r
            for r in ws.iter_rows(min_row=2)}
    n1 = rows[("StrainBook_0", "N1")]
    assert n1[2].value == "V"
    assert n1[3].value == pytest.approx(1.0 / RATE_SB)
    assert n1[4].value == blocks["StrainBook_0"]["N1"].size
    assert n1[5].value == T_START.isoformat()
    alpha = rows[("Positioner", "Alpha")]
    assert alpha[2].value == "deg"
    assert alpha[3].value == pytest.approx(1.0 / RATE_POS)
    # the synthesized time axis is described too
    t = rows[("Time", "Time")]
    assert t[2].value == "s"
    assert t[3].value == pytest.approx(1.0 / RATE_SB)


# ── Devices + Config sheets ──────────────────────────────────────────────
def test_devices_sheet_keeps_cal_pointers(workbook):
    _, wb, _ = workbook
    rows = {(r[0].value, r[1].value): r[2].value
            for r in wb["Devices"].iter_rows(min_row=2)}
    assert rows[("strainbook", "model")] == "StrainBook/616"
    assert rows[("strainbook", "cal_file")] == "C:/cals/sb616_2026.cal"
    assert rows[("daqbook", "cal_file")] == "C:/cals/db2000.pcf"


def test_config_sheet_json_round_trips(workbook):
    _, wb, _ = workbook
    lines = [r[0].value or "" for r in wb["Config"].iter_rows()]
    assert json.loads("\n".join(lines)) == CONFIG_SNAP


# ── sheet-name sanitization ──────────────────────────────────────────────
def test_sheet_name_sanitizer():
    used = set()
    assert sheet_name("StrainBook_0") == "StrainBook_0"
    assert sheet_name("a/b [c]: *?") == "a_b _c__ __"
    assert sheet_name("") == "Sheet"
    assert len(sheet_name("x" * 60)) == 31
    # case-insensitive dedup
    assert sheet_name("Tunnel", used) == "Tunnel"
    assert sheet_name("tunnel", used) == "tunnel_2"


def test_illegal_group_names_get_legal_sheets(tmp_path):
    rec = Hdf5Recorder(tmp_path, config_name="cfgY",
                       output_format="xlsx")
    long_name = "Group/with:illegal*chars?" + "x" * 30
    p = rec.write_point(
        point_meta={"alpha": 0.0},
        blocks={long_name: {"C1": np.arange(5.0)},
                "Meta": {"M1": np.arange(5.0)}},   # collides with fixed
        rates={long_name: 10.0, "Meta": 10.0})
    wb = load_workbook(p)
    assert all(len(n) <= 31 and not set("[]:*?/\\") & set(n)
               for n in wb.sheetnames)
    # the group named "Meta" must NOT shadow the fixed Meta sheet
    meta_keys = [r[0].value for r in wb["Meta"].iter_rows(min_row=2)]
    assert "alpha" in meta_keys
    # both data groups still land under their ORIGINAL names in Channels
    groups = {r[0].value for r in wb["Channels"].iter_rows(min_row=2)}
    assert long_name in groups and "Meta" in groups


# ── end-to-end: sim sweep per selectable format ──────────────────────────
FAKES = {
    "balance": {"adapter": "freestream._fakes.FakeStreamer",
                "enabled": True},
    "daq": {"adapter": "freestream._fakes.FakeDaq", "enabled": True},
    "pos": {"adapter": "freestream._fakes.FakePositioner",
            "enabled": True},
    "tun": {"adapter": "freestream._fakes.FakeTunnel", "enabled": True},
}
MODES = {"mode1": {"positioner": "pos", "balance": "balance",
                   "tunnel_conditions": "daq", "tunnel": "tun"}}


@pytest.mark.parametrize("fmt,ext", [("h5", ".h5"), ("mat", ".mat"),
                                     ("xlsx", ".xlsx")])
def test_sim_sweep_each_format_single_file_type(tmp_path, fmt, ext):
    if fmt == "mat":
        pytest.importorskip("scipy.io")
    from freestream.config import FreestreamConfig
    from freestream.manager import DeviceManager
    from freestream.runsheet import build_grid
    from freestream.sweep import DONE, SweepEngine

    manifest = tmp_path / "manifest.json"
    manifest.write_text(json.dumps({"modes": MODES, "devices": FAKES}),
                        encoding="utf-8")
    mgr = DeviceManager("mode1", sim=True, manifest_path=manifest)
    mgr.connect_all()
    for s in mgr.streaming:
        s.start()
    try:
        cfg = FreestreamConfig(samples=50, dwell_s=0.05, move_timeout_s=5,
                               tunnel_timeout_s=5, operator="pytest",
                               output_format=fmt)
        rec = Hdf5Recorder(tmp_path / "runs", config_name="testcfg",
                           output_format=cfg.output_format)
        engine = SweepEngine(mgr, rec, cfg)
        points = build_grid(alpha_spec="0", dwell_s=0.05, samples=50)
        outcomes = engine.run(points)
        assert [o.status for o in outcomes] == [DONE] * len(outcomes)
        p = Path(outcomes[0].path)
        assert p.suffix == ext and p.exists() and p.stat().st_size > 0
        # exactly ONE file, of exactly the selected type
        written = list((tmp_path / "runs").rglob("run_*.*"))
        assert [f.suffix for f in written] == [ext]
    finally:
        for s in mgr.streaming:
            s.stop()
        mgr.disconnect_all()
