"""Run-sheet import and sweep-grid expansion for Freestream (spec §5.2, §6.1).

A *run sheet* is a user-authored ``.csv``/``.xlsx`` spreadsheet with flexible,
user-defined columns.  Recognized columns (``alpha``, ``beta``, ``mach``,
``dwell_s``/``dwell``, ``samples``, ``air_state`` — case-insensitive) map onto
:class:`SweepPoint` fields; every OTHER column is inherited **verbatim** into
each expanded point's :attr:`SweepPoint.meta` dict.  That metadata-inheritance
payload is later stamped into the per-point HDF5 attrs, so hysteresis and
configuration variations (``flap_deg``, ``gear``, ``config_name``, operator,
notes, ...) are tracked without manual re-entry.  Unrecognized columns never
error — they are the whole point.

The tunnel axis is **Mach** (the MachLoop converts it to a fan-RPM command
at run time).  An ``rpm`` column is still accepted as a documented
*direct-RPM override*: its value is parsed to a float and stored in
``meta["rpm"]``; the sweep engine commands that RPM verbatim and bypasses
the Mach loop for the row (useful for rig commissioning before
``rpm_per_mach`` is tuned).

Axis cells use the ONE canonical sweep grammar in
:mod:`freestream.sweepgrammar` — a single number, a comma list (``0,2,4``), a
``start:delta:end`` range (``-4:2:8``; the MIDDLE value is the step), an ``R``
return-sweep suffix (``0:2:10R``), a ``@named`` reference, or ``csv:file``.
:func:`parse_axis_spec` here is a thin adapter over that grammar so the whole
suite shares exactly one expansion implementation.

Pure stdlib + openpyxl.  No Qt imports.
"""

from __future__ import annotations

import csv
import itertools
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple, Union

from . import sweepgrammar

__all__ = [
    "SweepPoint",
    "parse_axis_spec",
    "build_grid",
    "load_runsheet",
    "points_summary",
    "stamp_leg",
    "annotate_directions",
    "DEFAULT_DWELL_S",
    "DEFAULT_SAMPLES",
]

DEFAULT_DWELL_S: float = 0.5
DEFAULT_SAMPLES: int = 100
#: outermost → innermost. Aero axes nest mach → beta → alpha; traverse
#: axes nest z → y → x (x varies fastest — one plane at a time).
DEFAULT_ORDER: Tuple[str, ...] = ("mach", "beta", "alpha", "z", "y", "x")

_AXIS_NAMES = ("alpha", "beta", "mach", "x", "y", "z")

#: Anything that build_grid accepts as an axis spec.
AxisSpec = Union[None, str, float, int, Sequence[float]]


@dataclass
class SweepPoint:
    """One executable test point, expanded from a sweep grid or run-sheet row.

    ``meta`` carries ALL unrecognized run-sheet columns verbatim — the
    metadata-inheritance payload written into the point's HDF5 attrs.
    """

    alpha: Optional[float] = None
    beta: Optional[float] = None
    #: tunnel axis — target Mach number (converted to a fan-RPM command by
    #: freestream.machloop). A direct-RPM override, when requested via a
    #: run-sheet ``rpm`` column, lives in ``meta["rpm"]`` instead.
    mach: Optional[float] = None
    #: traverse axes [in] — Mode 3 probe-positioning sweeps (the traverse
    #: is the Positioner and the point matrix is spatial, not attitude)
    x: Optional[float] = None
    y: Optional[float] = None
    z: Optional[float] = None
    dwell_s: float = DEFAULT_DWELL_S
    samples: int = DEFAULT_SAMPLES
    air_state: str = "AirOn"
    meta: dict = field(default_factory=dict)
    row_index: int = 0
    status: str = "queued"
    #: "up" | "dn" | "" — alpha-dot sign, set only on hysteresis (return)
    #: sweeps so the two legs are distinguishable in filenames/metadata.
    direction: str = ""


# ---------------------------------------------------------------------------
# Axis-spec parsing
# ---------------------------------------------------------------------------

def parse_axis_spec(text: Union[str, float, int],
                    named: Optional[dict] = None) -> List[float]:
    """Parse one axis cell into an explicit list of setpoints.

    Thin adapter over :func:`freestream.sweepgrammar.expand` — the ONE
    canonical grammar (``start:delta:end`` ranges, ``R`` return sweeps,
    comma lists, ``@named`` references).  Accepted forms:

    * single number:          ``"5"``          → ``[5.0]``
    * explicit comma list:    ``"0,2,4"``      → ``[0.0, 2.0, 4.0]``
    * ``start:delta:end`` range (MIDDLE = step, end inclusive):
                              ``"-4:2:8"``      → ``[-4,-2,0,2,4,6,8]``
                              ``"10:-2:0"``     → ``[10, 8, 6, 4, 2, 0]``
    * trailing ``R`` = return sweep (hysteresis), mirrored back to the
      start dropping the apex: ``"0:2:10R"`` →
      ``[0, 2, 4, 6, 8, 10, 8, 6, 4, 2, 0]``.

    Raises ``ValueError`` for a blank/empty spec (callers pass ``None`` to
    omit an axis instead).  ``named`` supplies any ``@name`` definitions.
    """
    if isinstance(text, bool):  # bool is an int subclass; reject explicitly
        raise ValueError(f"invalid axis spec: {text!r}")
    if isinstance(text, (int, float)):
        return [float(text)]
    if not str(text).strip():
        raise ValueError("empty axis spec")
    values = sweepgrammar.expand(text, named=named)
    if not values:
        raise ValueError(f"invalid axis spec: {text!r}")
    return [float(v) for v in values]


# ---------------------------------------------------------------------------
# Grid expansion
# ---------------------------------------------------------------------------

def _axis_values(spec: AxisSpec,
                 named: Optional[dict] = None) -> List[Optional[float]]:
    """Normalize any accepted spec form to an explicit value list.

    ``None`` means the axis is omitted → a single ``None`` placeholder so the
    nested product still yields points.
    """
    if spec is None:
        return [None]
    if isinstance(spec, str) or isinstance(spec, (int, float)):
        return list(parse_axis_spec(spec, named=named))
    return [float(v) for v in spec]


def build_grid(
    alpha_spec: AxisSpec = None,
    beta_spec: AxisSpec = None,
    mach_spec: AxisSpec = None,
    x_spec: AxisSpec = None,
    y_spec: AxisSpec = None,
    z_spec: AxisSpec = None,
    order: Sequence[str] = DEFAULT_ORDER,
    dwell_s: float = DEFAULT_DWELL_S,
    samples: int = DEFAULT_SAMPLES,
    air_state: str = "AirOn",
    meta: Optional[dict] = None,
    row_index: int = 0,
    named: Optional[dict] = None,
) -> List[SweepPoint]:
    """Expand axis specs into a nested sweep grid of :class:`SweepPoint`.

    ``order`` lists axis names outermost-first (default: mach outer, beta,
    alpha innermost; traverse z outer, y, x innermost).  ``None`` specs mean
    that axis is omitted (its field stays ``None`` on every point).
    ``dwell_s``/``samples``/``air_state``/``meta``/``row_index`` are stamped
    onto every generated point; each point receives its own copy of ``meta``.
    ``named`` supplies ``@name`` definitions for the sweep grammar.
    """
    unknown = [name for name in order if name not in _AXIS_NAMES]
    if unknown:
        raise ValueError(
            f"unknown axis name(s) in order: {unknown!r}; expected {_AXIS_NAMES}"
        )
    specs = {"alpha": alpha_spec, "beta": beta_spec, "mach": mach_spec,
             "x": x_spec, "y": y_spec, "z": z_spec}
    values = {name: _axis_values(spec, named) for name, spec in specs.items()}

    # De-dup order, then append any axis it omits (innermost).
    loop_order: List[str] = []
    for name in tuple(order) + _AXIS_NAMES:
        if name not in loop_order:
            loop_order.append(name)

    points: List[SweepPoint] = []
    for combo in itertools.product(*(values[name] for name in loop_order)):
        assigned = dict(zip(loop_order, combo))
        points.append(
            SweepPoint(
                alpha=assigned["alpha"],
                beta=assigned["beta"],
                mach=assigned["mach"],
                x=assigned["x"],
                y=assigned["y"],
                z=assigned["z"],
                dwell_s=float(dwell_s),
                samples=int(samples),
                air_state=air_state,
                meta=dict(meta or {}),
                row_index=row_index,
            )
        )
    annotate_directions(points)
    return points


# ---------------------------------------------------------------------------
# Hysteresis: alpha-dot sign / sweep-direction tagging
# ---------------------------------------------------------------------------

def _set_direction(point: SweepPoint, direction: str) -> None:
    point.direction = direction
    # only stamp metadata on genuine hysteresis legs — monotonic sweeps
    # keep their meta untouched (no spurious columns in the HDF5 attrs).
    if direction:
        point.meta["sweep_dir"] = direction
        point.meta["alpha_dot"] = 1 if direction == "up" else -1


def stamp_leg(point: SweepPoint, leg: Optional[str]) -> None:
    """Public leg tagger used by the run-sheet workbook expansion.

    ``leg`` is the ``"up"`` / ``"dn"`` return-sweep tag from
    :func:`freestream.sweepgrammar.build_points` (``None`` on a monotonic
    axis).  Rides in ``.direction`` and — for genuine legs — in ``.meta``
    (``sweep_dir`` / ``alpha_dot``) exactly like :func:`build_grid`, so the
    two hysteresis legs never collide in the filename or HDF5 attrs.
    """
    _set_direction(point, leg or "")


def _label_leg(leg: List[SweepPoint]) -> None:
    """Tag one α sweep (fixed beta/mach) with up/dn IFF it reverses.

    Monotonic sweeps get no tag (clean filenames); only a return sweep
    (alpha ascending then descending, or vice-versa) — where the same
    alpha is visited with opposite α̇ — gets its two legs labelled so the
    hysteresis points never collide in the filename or metadata.
    """
    alphas = [p.alpha for p in leg]
    if len(leg) < 2 or any(a is None for a in alphas):
        for p in leg:
            _set_direction(p, "")
        return
    deltas = [alphas[k + 1] - alphas[k] for k in range(len(alphas) - 1)]
    has_up = any(d > 0 for d in deltas)
    has_dn = any(d < 0 for d in deltas)
    if not (has_up and has_dn):
        for p in leg:                       # monotonic — no direction tag
            _set_direction(p, "")
        return
    for k, p in enumerate(leg):
        d = deltas[0] if k == 0 else deltas[k - 1]
        _set_direction(p, "up" if d > 0 else "dn" if d < 0 else "")


def annotate_directions(points: List[SweepPoint]) -> None:
    """Split the point list into consecutive fixed-(beta, mach) alpha legs
    and label each leg's hysteresis direction in place."""
    i, n = 0, len(points)
    while i < n:
        key = (points[i].beta, points[i].mach)
        j = i
        while j < n and (points[j].beta, points[j].mach) == key:
            j += 1
        _label_leg(points[i:j])
        i = j


# ---------------------------------------------------------------------------
# Run-sheet loading (.csv / .xlsx)
# ---------------------------------------------------------------------------

def _canon(header: object) -> str:
    return str(header).strip().lower().replace(" ", "_").replace("-", "_")


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _normalize_air_state(value: object) -> str:
    key = str(value).strip().lower().replace(" ", "").replace("_", "")
    if key in ("airon", "on", "wind", "windon", "1", "true"):
        return "AirOn"
    if key in ("airoff", "off", "windoff", "tare", "0", "false"):
        return "AirOff"
    return str(value).strip()  # pass unknown labels through verbatim


def _read_csv_rows(path: Path) -> Iterable[List[Tuple[str, object]]]:
    """Yield each data row as ordered (original_header, cell_value) pairs."""
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        try:
            headers = next(reader)
        except StopIteration:
            return
        for raw in reader:
            raw = list(raw) + [""] * (len(headers) - len(raw))
            yield [
                (header, raw[i])
                for i, header in enumerate(headers)
                if str(header).strip()
            ]


def _read_xlsx_rows(path: Path) -> Iterable[List[Tuple[str, object]]]:
    """Yield each data row of the first worksheet as (header, value) pairs."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required to read .xlsx run sheets "
            "(pip install openpyxl)"
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            return
        for raw in rows:
            raw = list(raw) + [None] * (len(headers) - len(raw))
            yield [
                (str(header), raw[i])
                for i, header in enumerate(headers)
                if header is not None and str(header).strip()
            ]
    finally:
        workbook.close()


def load_runsheet(path: Union[str, Path]) -> List[SweepPoint]:
    """Load a ``.csv``/``.xlsx`` run sheet into an expanded point list.

    Column handling is flexible and case-insensitive.  Recognized columns:
    ``alpha``, ``beta``, ``mach`` (single numbers OR axis specs — a spec cell
    expands the row into multiple points via :func:`build_grid`),
    ``dwell_s``/``dwell``, ``samples``, ``air_state`` (default ``AirOn``),
    and ``rpm`` — a documented direct-RPM override parsed to a float and
    stored in ``meta["rpm"]`` (the sweep engine commands it verbatim,
    bypassing the Mach loop for those points).
    ALL other columns are inherited verbatim into each point's ``.meta`` —
    unrecognized columns never error.  Fully blank rows are skipped.
    ``row_index`` on each point is the 0-based data-row index in the sheet.
    """
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv_rows(path)
    elif suffix in (".xlsx", ".xlsm"):
        rows = _read_xlsx_rows(path)
    else:
        raise ValueError(
            f"unsupported run-sheet type {suffix!r} for {path} "
            "(expected .csv or .xlsx)"
        )

    points: List[SweepPoint] = []
    for row_index, pairs in enumerate(rows):
        if all(_is_blank(value) for _, value in pairs):
            continue  # skip fully blank rows

        specs: dict = {name: None for name in _AXIS_NAMES}
        dwell_s = DEFAULT_DWELL_S
        samples = DEFAULT_SAMPLES
        air_state = "AirOn"
        meta: dict = {}

        for header, value in pairs:
            key = _canon(header)
            if key in _AXIS_NAMES:
                if not _is_blank(value):
                    specs[key] = value  # number or axis-spec string
            elif key == "rpm":
                # documented direct-RPM override → meta (bypasses MachLoop)
                if not _is_blank(value):
                    meta["rpm"] = float(str(value).strip())
            elif key in ("dwell_s", "dwell"):
                if not _is_blank(value):
                    dwell_s = float(str(value).strip())
            elif key == "samples":
                if not _is_blank(value):
                    samples = int(float(str(value).strip()))
            elif key in ("air_state", "airstate", "air"):
                if not _is_blank(value):
                    air_state = _normalize_air_state(value)
            else:
                meta[header] = value  # metadata inheritance — verbatim

        points.extend(
            build_grid(
                alpha_spec=specs["alpha"],
                beta_spec=specs["beta"],
                mach_spec=specs["mach"],
                x_spec=specs["x"],
                y_spec=specs["y"],
                z_spec=specs["z"],
                dwell_s=dwell_s,
                samples=samples,
                air_state=air_state,
                meta=meta,
                row_index=row_index,
            )
        )
    return points


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

def _fmt(value: float) -> str:
    return f"{value:g}"


def points_summary(points: Sequence[SweepPoint]) -> str:
    """Short human summary, e.g. ``"12 points: alpha -4..8, 2 mach levels, 3 rows"``."""
    n = len(points)
    if n == 0:
        return "0 points"
    parts: List[str] = []
    for axis in ("alpha", "beta", "x", "y", "z"):
        vals = sorted({getattr(p, axis) for p in points
                       if getattr(p, axis) is not None})
        if not vals:
            continue
        if len(vals) == 1:
            parts.append(f"{axis} {_fmt(vals[0])}")
        else:
            parts.append(f"{axis} {_fmt(vals[0])}..{_fmt(vals[-1])}")
    mach_vals = sorted({p.mach for p in points if p.mach is not None})
    if len(mach_vals) == 1:
        parts.append(f"mach {_fmt(mach_vals[0])}")
    elif mach_vals:
        parts.append(f"{len(mach_vals)} mach levels")
    n_rows = len({p.row_index for p in points})
    parts.append(f"{n_rows} row" + ("s" if n_rows != 1 else ""))
    return f"{n} point" + ("s" if n != 1 else "") + ": " + ", ".join(parts)
