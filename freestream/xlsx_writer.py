"""Excel (.xlsx) per-point output — the spreadsheet-review format.

One workbook per test point (same ``run_NNNN_…`` basename as the HDF5
schema), written via openpyxl from the recorder's in-memory blocks — no
intermediate ``.h5``. Unlike HDF5/.mat this format is NOT readable by
Streamlined for reduction; it exists so an engineer can eyeball a point
in Excel.

Workbook layout
---------------
* one **Data sheet per HDF5 group** (``StrainBook_0``, ``DaqBook2005``,
  ``Positioner``, ``Tunnel``, …): column A is ``Time`` in seconds built
  from that group's sample rate (``i / rate``), then one column per
  channel. Row 1 = channel names, row 2 = units, data from row 3.
  Numbers are written as numbers, never strings.
* **Meta** — key/value rows: EVERY root attr verbatim (run params, test
  info, ref dims, air_state, mode, operator, custom run-sheet columns…).
* **Channels** — one row per channel: group, channel, unit,
  wf_increment, wf_samples, wf_start_time (incl. the synthesized Time
  axis when present).
* **Devices** — device / key / value rows of the per-device metadata
  (cal_file stays a POINTER string, never applied data).
* **Config** — the measurement-config snapshot JSON, pretty-printed one
  line per row.

Sheet names are made xlsx-legal by :func:`sheet_name`: the characters
``[ ] : * ? / \\`` become ``_``, names are trimmed to Excel's 31-char
limit, never empty, never start/end with an apostrophe, and collide-safe
(case-insensitive) via ``_2``, ``_3`` … suffixes. The fixed Meta /
Channels / Devices / Config names are reserved up front so a data group
with one of those names cannot shadow them.
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Set, Union

import numpy as np

__all__ = ["write_xlsx", "sheet_name"]

# characters Excel forbids in sheet names
_BAD_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_SHEET_NAME_MAX = 31          # Excel hard limit


def sheet_name(name: Any, used: Optional[Set[str]] = None) -> str:
    """Sanitize *name* into a legal, unique Excel sheet name.

    ``[ ] : * ? / \\`` → ``_``; trimmed to 31 chars; never empty and
    never starts/ends with ``'``. When *used* is given (a set of
    LOWER-CASED taken names — Excel sheet names are case-insensitive),
    collisions get ``_2``, ``_3`` … suffixes and the result is added.
    """
    s = _BAD_SHEET_CHARS.sub("_", str(name)).strip().strip("'").strip()
    s = s[:_SHEET_NAME_MAX] or "Sheet"
    if used is not None:
        base, i = s, 2
        while s.lower() in used:
            suffix = f"_{i}"
            s = base[:_SHEET_NAME_MAX - len(suffix)] + suffix
            i += 1
        used.add(s.lower())
    return s


def _cell(value: Any) -> Any:
    """openpyxl-friendly rendering of one attr value (mirrors the HDF5
    ``_set_attr`` semantics): str/bool verbatim, numbers as NUMBERS
    (numpy scalars unwrapped), datetime/Path stringified, dict → JSON,
    lists/arrays → JSON so the cell stays self-describing."""
    if value is None:
        return None
    if isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return json.dumps(value)
    if isinstance(value, (list, tuple, np.ndarray)):
        arr = np.asarray(value)
        return json.dumps(arr.tolist() if arr.dtype.kind not in ("U", "S",
                                                                 "O")
                          else [str(v) for v in value])
    return str(value)


def write_xlsx(
    path: Union[str, Path],
    root_attrs: Mapping[str, Any],
    blocks: Dict[str, Dict[str, np.ndarray]],
    rates: Dict[str, float],
    units: Dict[tuple, str],
    start_iso: str,
    device_meta: Optional[List[Dict[str, Any]]] = None,
    config_snapshot: Optional[Dict[str, Any]] = None,
    time_axis: Optional[np.ndarray] = None,
) -> Path:
    """Write one test point as an Excel workbook and return its path.

    Parameters mirror ``Hdf5Recorder._write_h5``: *blocks* is
    ``{group: {channel: 1-D raw array}}`` (verbatim float64, NO
    calibration applied), *rates* the per-group sample rates, *units* the
    normalized ``{(group|None, channel): unit}`` map, *time_axis* the
    synthesized seconds axis (None when the caller supplied a "Time"
    block — it then appears as a regular group).
    """
    from openpyxl import Workbook

    path = Path(path)
    wb = Workbook()
    wb.remove(wb.active)                        # drop the default sheet
    # reserve the fixed sheet names so a data group can never shadow them
    used: Set[str] = {"meta", "channels", "devices", "config"}

    def unit_of(group: str, ch: str) -> str:
        return (units.get((group, ch)) or units.get((None, ch)) or "")

    # ── one Data sheet per group ─────────────────────────────────────────
    chan_rows: List[tuple] = []                 # feeds the Channels sheet
    for group_name, channels in blocks.items():
        ws = wb.create_sheet(sheet_name(group_name, used))
        rate = float(rates.get(group_name, 0.0))
        incr = (1.0 / rate) if rate > 0 else 0.0
        arrays = {str(ch): np.asarray(data, dtype=np.float64).ravel()
                  for ch, data in channels.items()}
        ws.append(["Time"] + list(arrays))                     # row 1: names
        ws.append(["s"] + [unit_of(group_name, ch) for ch in arrays])
        n = max((a.size for a in arrays.values()), default=0)
        cols = list(arrays.values())
        for i in range(n):
            ws.append([i * incr] + [float(a[i]) if i < a.size else None
                                    for a in cols])
        ws.freeze_panes = "A3"
        for ch, arr in arrays.items():
            chan_rows.append((str(group_name), ch, unit_of(group_name, ch),
                              incr, int(arr.size), start_iso))

    # synthesized Time axis → Channels row only (every Data sheet already
    # carries its own Time column; a dedicated sheet would be redundant)
    if time_axis is not None:
        t = np.asarray(time_axis, dtype=np.float64).ravel()
        t_incr = float(t[1] - t[0]) if t.size > 1 else 0.0
        chan_rows.append(("Time", "Time", "s", t_incr, int(t.size),
                          start_iso))

    # ── Meta: every root attr verbatim, key/value rows ───────────────────
    ws = wb.create_sheet("Meta")
    ws.append(["key", "value"])
    for key, val in root_attrs.items():
        cell = _cell(val)
        if cell is None:                        # None skipped, like _set_attr
            continue
        ws.append([str(key), cell])
    ws.freeze_panes = "A2"

    # ── Channels: per-channel waveform attrs ─────────────────────────────
    ws = wb.create_sheet("Channels")
    ws.append(["group", "channel", "unit", "wf_increment", "wf_samples",
               "wf_start_time"])
    for row in chan_rows:
        ws.append(list(row))
    ws.freeze_panes = "A2"

    # ── Devices: per-device metadata (cal_file stays a POINTER) ─────────
    ws = wb.create_sheet("Devices")
    ws.append(["device", "key", "value"])
    for i, dev in enumerate(device_meta or []):
        name = str(dev.get("id") or dev.get("name") or dev.get("model")
                   or f"device_{i}")
        for key, val in dev.items():
            cell = _cell(val)
            if cell is None:
                continue
            ws.append([name, str(key), cell])
    ws.freeze_panes = "A2"

    # ── Config: the measurement-config snapshot, pretty JSON ────────────
    ws = wb.create_sheet("Config")
    snapshot = config_snapshot if config_snapshot is not None else {}
    for line in json.dumps(snapshot, indent=2).splitlines():
        ws.append([line])

    wb.save(path)
    return path
