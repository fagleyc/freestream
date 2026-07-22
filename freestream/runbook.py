"""Run-sheet WORKBOOK loader — Casey's 5-sheet standardized template (§5).

A *run book* is the ``Freestream_RunSheet_Template.xlsx`` format: five sheets
(``Guide``, ``Test Info``, ``Run Matrix``, ``Model Configs``, ``Named
Arrays``).  :func:`load_runbook` parses it into a :class:`RunBook` dataclass,
locating header rows by their LABEL TEXT (robust to inserted rows or extra
columns), and :func:`build_run_points` expands one selected run row into the
executable :class:`~freestream.runsheet.SweepPoint` matrix using the ONE sweep
grammar (:mod:`freestream.sweepgrammar`).

The older flat single-sheet CSV/XLSX path
(:func:`freestream.runsheet.load_runsheet`) still works as a simple fallback;
the workbook is the primary path.

Pure Python (stdlib + openpyxl).  No Qt imports.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple, Union

from . import sweepgrammar
from .runsheet import (DEFAULT_DWELL_S, DEFAULT_SAMPLES, SweepPoint,
                       stamp_leg)

__all__ = [
    "RunRow",
    "RunBook",
    "load_runbook",
    "build_run_points",
    "expanded_count",
    "is_runbook_workbook",
]

# Sheet names (matched case-insensitively / loosely so a lightly renamed
# tab still resolves).
_SHEET_ALIASES = {
    "test_info": ("test info", "testinfo", "test"),
    "run_matrix": ("run matrix", "runmatrix", "matrix", "runs"),
    "model_configs": ("model configs", "modelconfigs", "configs", "config"),
    "named_arrays": ("named arrays", "namedarrays", "arrays", "named"),
}

# Test-Info friendly-key aliases (normalized label → canonical field). Each
# value lists candidate normalized-label fragments, highest priority first.
_INFO_ALIASES: Dict[str, Sequence[str]] = {
    "test_name": ("test_entry_name", "entry_name", "test_name"),
    "facility": ("facility",),
    "model_name": ("model_name_no", "model_name", "model"),
    "engineer": ("test_engineer", "engineer"),
    "operator": ("operator_s", "operators", "operator"),
    "start_date": ("start_date",),
    "end_date": ("end_date",),
    "freestream_version": ("freestream_version", "version"),
    "objectives": ("test_objectives", "objectives"),
    "balance_id": ("balance_id", "balance"),
    "sting_config": ("sting_mount_config", "sting"),
    "nominal_mach": ("nominal_mach_range", "nominal_mach"),
    "nominal_re": ("nominal_re",),
    "data_prefix": ("data_file_prefix", "data_prefix", "prefix"),
    "blockage": ("blockage_wall_corr", "blockage"),
    "notes": ("notes",),
}

_REF_SYMBOLS = ("Sref", "cref", "bref", "MRC_x", "MRC_y", "MRC_z")


# ── dataclasses ────────────────────────────────────────────────────────────
@dataclass
class RunRow:
    """One row of the Run Matrix (a single named run, pre-expansion)."""

    run: str
    enable: bool
    alpha_cell: str
    beta_cell: str
    mach_cell: str
    samples: Optional[int]
    sample_rate_hz: Optional[float]
    config_name: str
    notes: str
    row_index: int          # 1-based worksheet row (traceability)


@dataclass
class RunBook:
    """A parsed run-sheet workbook."""

    test_info: Dict[str, object] = field(default_factory=dict)
    #: {symbol -> {"quantity","symbol","value","units","notes"}}
    ref_dims: Dict[str, dict] = field(default_factory=dict)
    #: {config_name -> {column_header -> value}} (model-specific cols verbatim)
    configs: Dict[str, dict] = field(default_factory=dict)
    #: {name -> definition-cell}
    named_arrays: Dict[str, str] = field(default_factory=dict)
    runs: List[RunRow] = field(default_factory=list)
    source_path: str = ""

    # -- convenience views ------------------------------------------------
    def friendly_info(self) -> Dict[str, object]:
        """Map the raw Test-Info labels onto canonical field names."""
        out: Dict[str, object] = {}
        for field_name, fragments in _INFO_ALIASES.items():
            value = self._lookup_info(fragments)
            if value is not None:
                out[field_name] = value
        return out

    def _lookup_info(self, fragments: Sequence[str]) -> Optional[object]:
        keys = list(self.test_info)
        for frag in fragments:                    # exact match wins
            if frag in self.test_info:
                return self.test_info[frag]
        for frag in fragments:                    # then substring
            for key in keys:
                if frag in key:
                    return self.test_info[key]
        return None

    def ref_dim_value(self, symbol: str) -> Optional[float]:
        entry = self.ref_dims.get(symbol)
        if not entry:
            return None
        value = entry.get("value")
        try:
            return float(value) if value is not None else None
        except (TypeError, ValueError):
            return None

    @property
    def enabled_runs(self) -> List[RunRow]:
        return [r for r in self.runs if r.enable]


# ── label / value helpers ──────────────────────────────────────────────────
def _norm(text: object) -> str:
    """Normalize a label to ``snake_case`` alnum key (drops punctuation)."""
    s = re.sub(r"[^0-9a-z]+", "_", str(text).strip().lower())
    return s.strip("_")


def _blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _as_bool(value: object, default: bool = True) -> bool:
    if _blank(value):
        return default
    if isinstance(value, bool):
        return value
    key = str(value).strip().lower()
    if key in ("y", "yes", "true", "1", "on", "enable", "enabled"):
        return True
    if key in ("n", "no", "false", "0", "off", "disable", "disabled"):
        return False
    return default


def _as_int(value: object) -> Optional[int]:
    if _blank(value):
        return None
    try:
        return int(round(float(str(value).strip())))
    except (TypeError, ValueError):
        return None


def _as_float(value: object) -> Optional[float]:
    if _blank(value):
        return None
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _cell_text(value: object) -> str:
    """Axis-cell text: numbers stringified, blanks → '' (empty spec)."""
    if _blank(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


# ── worksheet access ───────────────────────────────────────────────────────
def _sheet_rows(workbook, aliases: Sequence[str]) -> List[tuple]:
    """Return a worksheet's rows (as value tuples) by fuzzy title match."""
    wanted = [a.replace(" ", "") for a in aliases]
    for ws in workbook.worksheets:
        title = ws.title.strip().lower().replace(" ", "")
        if any(title == w or w in title for w in wanted):
            return [tuple(r) for r in ws.iter_rows(values_only=True)]
    return []


def _get(row: tuple, idx: int) -> object:
    return row[idx] if 0 <= idx < len(row) else None


def _find_header_row(rows: Sequence[tuple], *required: str
                     ) -> Tuple[int, Dict[str, int]]:
    """Locate the header row containing every ``required`` fragment.

    Returns ``(row_index, {normalized_header -> column_index})`` or
    ``(-1, {})`` when not found.
    """
    for i, row in enumerate(rows):
        norms = {}
        for col, cell in enumerate(row):
            if not _blank(cell):
                norms.setdefault(_norm(cell), col)
        if all(any(req in key for key in norms) for req in required):
            return i, norms
    return -1, {}


# ── Test Info ───────────────────────────────────────────────────────────────
def _parse_test_info(rows: Sequence[tuple]
                     ) -> Tuple[Dict[str, object], Dict[str, dict]]:
    """Parse the Test Info tab into (header fields, reference dimensions)."""
    info: Dict[str, object] = {}
    ref_dims: Dict[str, dict] = {}

    # locate the reference-dimensions block so header-field scanning stops
    ref_title = next((i for i, r in enumerate(rows)
                      if any("reference_dimension" in _norm(c)
                             for c in r if not _blank(c))), len(rows))

    # header fields: two label/value blocks per row — (B,C) and (E,F). We
    # scan every row above the ref-dims block and treat any string cell
    # immediately followed by a value cell as a label/value pair.
    for row in rows[:ref_title]:
        for label_col, value_col in ((1, 2), (4, 5)):
            label = _get(row, label_col)
            if _blank(label) or not isinstance(label, str):
                continue
            key = _norm(label)
            if not key or key in ("test_entry_information",):
                continue
            info[key] = _get(row, value_col)

    # reference dimensions table
    hdr, cols = _find_header_row(rows[ref_title:], "symbol", "value")
    if hdr >= 0:
        base = ref_title + hdr
        c_qty = cols.get("quantity")
        c_sym = cols.get("symbol")
        c_val = cols.get("value")
        c_unit = cols.get("units")
        c_note = cols.get("notes")
        for row in rows[base + 1:]:
            symbol = _get(row, c_sym) if c_sym is not None else None
            if _blank(symbol):
                continue
            sym = str(symbol).strip()
            ref_dims[sym] = {
                "quantity": _get(row, c_qty) if c_qty is not None else None,
                "symbol": sym,
                "value": _as_float(_get(row, c_val))
                if c_val is not None else None,
                "units": _get(row, c_unit) if c_unit is not None else None,
                "notes": _get(row, c_note) if c_note is not None else None,
            }
    return info, ref_dims


# ── Run Matrix ──────────────────────────────────────────────────────────────
def _resolve_matrix_cols(norms: Dict[str, int]) -> Dict[str, int]:
    """Map Run-Matrix fields to their column indices (header text may vary)."""
    out: Dict[str, int] = {}
    for key, col in norms.items():
        if key == "run" and "run" not in out:
            out["run"] = col
        elif key.startswith("enable"):
            out["enable"] = col
        elif "alpha" in key:
            out["alpha"] = col
        elif "beta" in key:
            out["beta"] = col
        elif key == "mach" or key.startswith("mach"):
            out.setdefault("mach", col)
        elif key == "samples":
            out["samples"] = col
        elif "sample_rate" in key or key == "rate" or key.endswith("_hz"):
            out["sample_rate"] = col
        elif key.startswith("config") and "config" not in out:
            out["config"] = col
        elif key.startswith("note"):
            out["notes"] = col
    return out


def _parse_run_matrix(rows: Sequence[tuple]) -> List[RunRow]:
    hdr, norms = _find_header_row(rows, "run", "mach")
    if hdr < 0:
        return []
    cols = _resolve_matrix_cols(norms)
    runs: List[RunRow] = []
    for offset, row in enumerate(rows[hdr + 1:], start=hdr + 2):  # 1-based row
        run_name = _get(row, cols.get("run", -1))
        if _blank(run_name):
            if all(_blank(c) for c in row):
                continue
            continue                              # a row with no run id
        runs.append(RunRow(
            run=str(run_name).strip(),
            enable=_as_bool(_get(row, cols.get("enable", -1))),
            alpha_cell=_cell_text(_get(row, cols.get("alpha", -1))),
            beta_cell=_cell_text(_get(row, cols.get("beta", -1))),
            mach_cell=_cell_text(_get(row, cols.get("mach", -1))),
            samples=_as_int(_get(row, cols.get("samples", -1))),
            sample_rate_hz=_as_float(_get(row, cols.get("sample_rate", -1))),
            config_name=_cell_text(_get(row, cols.get("config", -1))),
            notes=_cell_text(_get(row, cols.get("notes", -1))),
            row_index=offset,
        ))
    return runs


# ── Model Configs / Named Arrays ────────────────────────────────────────────
def _parse_configs(rows: Sequence[tuple]) -> Dict[str, dict]:
    hdr, _ = _find_header_row(rows, "config_name")
    if hdr < 0:
        return {}
    header_row = rows[hdr]
    # column headers verbatim (skip blank/placeholder "add columns" cells)
    headers: List[Tuple[int, str]] = []
    name_col = 0
    for col, cell in enumerate(header_row):
        if _blank(cell):
            continue
        text = str(cell).strip()
        if "add column" in text.lower():
            continue
        headers.append((col, text))
        if _norm(text) == "config_name":
            name_col = col
    configs: Dict[str, dict] = {}
    for row in rows[hdr + 1:]:
        name = _get(row, name_col)
        if _blank(name):
            continue
        record = {text: _get(row, col) for col, text in headers}
        configs[str(name).strip()] = record
    return configs


def _parse_named_arrays(rows: Sequence[tuple]) -> Dict[str, str]:
    hdr, norms = _find_header_row(rows, "name", "definition")
    if hdr < 0:
        return {}
    name_col = norms.get("name", 0)
    def_col = norms.get("definition", 1)
    out: Dict[str, str] = {}
    for row in rows[hdr + 1:]:
        name = _get(row, name_col)
        definition = _get(row, def_col)
        if _blank(name) or _blank(definition):
            continue
        out[str(name).strip()] = str(definition).strip()
    return out


# ── public loader ──────────────────────────────────────────────────────────
def is_runbook_workbook(path: Union[str, Path]) -> bool:
    """True when *path* looks like the 5-sheet run-sheet workbook (has a
    ``Run Matrix`` tab), as opposed to a flat single-sheet run sheet."""
    path = Path(path)
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return False
    try:
        from openpyxl import load_workbook
    except ImportError:                            # pragma: no cover
        return False
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:                              # noqa: BLE001
        return False
    try:
        titles = [t.strip().lower().replace(" ", "") for t in wb.sheetnames]
        return any("runmatrix" in t or t == "matrix" for t in titles)
    finally:
        wb.close()


def load_runbook(path: Union[str, Path]) -> RunBook:
    """Parse a run-sheet workbook into a :class:`RunBook`."""
    path = Path(path)
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(
            f"run book must be an .xlsx workbook, got {path.suffix!r} "
            f"(use load_runsheet for flat CSV/single-sheet sheets)")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:                     # pragma: no cover
        raise ImportError("openpyxl is required to read run-sheet workbooks "
                          "(pip install openpyxl)") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        info, ref_dims = _parse_test_info(
            _sheet_rows(wb, _SHEET_ALIASES["test_info"]))
        runs = _parse_run_matrix(
            _sheet_rows(wb, _SHEET_ALIASES["run_matrix"]))
        configs = _parse_configs(
            _sheet_rows(wb, _SHEET_ALIASES["model_configs"]))
        named = _parse_named_arrays(
            _sheet_rows(wb, _SHEET_ALIASES["named_arrays"]))
    finally:
        wb.close()
    return RunBook(test_info=info, ref_dims=ref_dims, configs=configs,
                   named_arrays=named, runs=runs, source_path=str(path))


# ── run expansion ──────────────────────────────────────────────────────────
def _base_meta(runbook: RunBook, run_row: RunRow) -> dict:
    meta: dict = {}
    # model-config columns verbatim (incl. model-specific ones)
    meta.update(runbook.configs.get(run_row.config_name, {}))
    meta["run"] = run_row.run
    meta["config_name"] = run_row.config_name
    # selected test-info fields
    info = runbook.friendly_info()
    for key in ("test_name", "model_name", "operator", "data_prefix",
                "facility", "engineer", "objectives"):
        value = info.get(key)
        if value not in (None, ""):
            meta[key] = value
    # reference dimensions (feed Streamlined's coefficient reduction)
    for symbol in _REF_SYMBOLS:
        value = runbook.ref_dim_value(symbol)
        if value is not None:
            meta[symbol] = value
    return meta


def build_run_points(runbook: RunBook, run_row: RunRow) -> List[SweepPoint]:
    """Expand one Run-Matrix row into executable :class:`SweepPoint` s.

    Mach is nested outermost (air-off ``0`` auto-prepended) → beta → alpha
    innermost, via :func:`freestream.sweepgrammar.build_points` using the
    run book's Named Arrays for ``@name`` references.  Each point's ``meta``
    carries the referenced config's columns verbatim, selected test-info
    fields, the reference dimensions, the run name, and the up/dn leg tag;
    ``samples`` comes from the row (overriding the global default).
    """
    specs = sweepgrammar.build_points(
        run_row.alpha_cell, run_row.beta_cell, run_row.mach_cell,
        named=runbook.named_arrays)
    base = _base_meta(runbook, run_row)
    samples = run_row.samples if run_row.samples is not None else DEFAULT_SAMPLES
    points: List[SweepPoint] = []
    for spec in specs:
        point = SweepPoint(
            alpha=spec["alpha"], beta=spec["beta"], mach=spec["mach"],
            dwell_s=DEFAULT_DWELL_S, samples=int(samples),
            meta=dict(base), row_index=run_row.row_index)
        stamp_leg(point, spec["leg"])
        points.append(point)
    return points


def expanded_count(runbook: RunBook, run_row: RunRow) -> int:
    """Number of test points ``run_row`` expands to (live matrix preview)."""
    return len(sweepgrammar.build_points(
        run_row.alpha_cell, run_row.beta_cell, run_row.mach_cell,
        named=runbook.named_arrays))
